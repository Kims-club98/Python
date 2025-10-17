from openpyxl import Workbook


#워크시트 시트창 하나 추가하고 색깔 추가
wb=Workbook()

#wb가 active되면
ws=wb.active

ws.title='기본시트'

ws=wb.create_sheet()

ws.title='나의시트'

#시트 색깔 지정하기
ws.sheet_properties.taColor="FF00DD"

ws1=wb.create_sheet('너의시트')
your_sheet=wb['너의시트']
your_sheet['A1']='테스트'

#your_sheet를 카피함
copy_sheet=wb.copy_worksheet(your_sheet)
copy_sheet.title='카피시트'

print(wb.sheetnames)
#
wb.save('data/sample1.xlsx')
wb.close()
