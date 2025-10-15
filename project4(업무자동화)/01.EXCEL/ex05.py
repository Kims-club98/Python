from openpyxl import load_workbook #엑셀파일을 불러오는 함수

wb=load_workbook('data/sample3.xlsx')
ws=wb.active

for row in ws.iter_rows(min_row=2, max_row=5): #행값을 하나씩 가지고 옴(min_row=2는 2행부터 가지고 옴)(min_row=2, max_row=5는 2행부터 5행까지 가져옴)
    for col in row: #행값에서 열을 가지고 옴
        print(col.value, end=',') #값을 출려해줌, 각 구분자는 ,로 구분
    print(' ') #행 기준(열기준X)으로 구분해줌    
