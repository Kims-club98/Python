from openpyxl import Workbook

wb=Workbook()
ws=wb.active
ws.title='나의 시트' #엑셀 활성화 후 기본시트를 나의시트로

#각 셀의 위치에 값 넣기
ws['A1']=1
ws['A2']=2
ws['B1']=3
ws['B2']=4


# print(ws['A1']) #ws['A1']은 셀의 정보를 출력
# print(ws['A1'].value) #ws['A1']의 셀값을 출력
# print(ws.cell(row=1,column=1).value)#셀 지정이 아닌 1행, 1열의 값을 출력

for row in range(1,3): #excel의 경우에는 1부터 시작하고 2까지 출력함(그 외는 0부터 시작)
    for col in range(1,3): #excel의 경우에는 1부터 시작하고 2까지 출력함(그 외는 0부터 시작)
        print(ws.cell(row=row,column=col).value)
        # 행은 1부터 3까지, 열은 1,2까지 하나씩 꺼내서 ws.cell로 출력하여 값만 나타냄(value)

wb.save('data/sample2.xlsx')
wb.close()